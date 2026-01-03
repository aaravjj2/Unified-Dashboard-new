"""
Export Service for PDF/Excel Generation
Implements #257 from ROADMAP_ULTIMATE.md
"""
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any, Union
from datetime import datetime
from io import BytesIO
import base64
import logging
import json

logger = logging.getLogger(__name__)

# Optional imports
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExportService:
    """
    Export data to PDF and Excel formats
    """
    
    def __init__(self):
        self.styles = None
        if HAS_REPORTLAB:
            self.styles = getSampleStyleSheet()
            self._setup_custom_styles()
            
    def _setup_custom_styles(self):
        """Setup custom PDF styles"""
        if not HAS_REPORTLAB:
            return
            
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#1a1a2e')
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.HexColor('#2d3436')
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=8
        ))
        
    def export_to_excel(self, data: Dict[str, Any], filename: str = None) -> Union[bytes, str]:
        """
        Export data to Excel format
        
        Args:
            data: Dictionary with sheets and their data
            filename: If provided, save to file; otherwise return bytes
            
        Returns:
            Bytes of Excel file or filepath
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")
            
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Style definitions
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        sheet_idx = 0
        for sheet_name, sheet_data in data.items():
            if sheet_idx == 0:
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            if isinstance(sheet_data, pd.DataFrame):
                self._write_dataframe_to_sheet(ws, sheet_data, header_font, header_fill, border)
            elif isinstance(sheet_data, dict):
                self._write_dict_to_sheet(ws, sheet_data, header_font, header_fill, border)
            elif isinstance(sheet_data, list):
                self._write_list_to_sheet(ws, sheet_data, header_font, header_fill, border)
                
            sheet_idx += 1
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        if filename:
            with open(filename, 'wb') as f:
                f.write(output.getvalue())
            return filename
        
        return output.getvalue()
    
    def _write_dataframe_to_sheet(self, ws, df, header_font, header_fill, border):
        """Write DataFrame to Excel sheet"""
        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=self._format_value(value))
                cell.border = border
                
                # Format numbers
                if isinstance(value, (int, float)) and not np.isnan(value) if isinstance(value, float) else True:
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(value, float):
                        cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
            
    def _write_dict_to_sheet(self, ws, data, header_font, header_fill, border):
        """Write dictionary to Excel sheet"""
        ws.cell(row=1, column=1, value='Key').font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value='Value').font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        
        for row_idx, (key, value) in enumerate(data.items(), 2):
            ws.cell(row=row_idx, column=1, value=str(key)).border = border
            ws.cell(row=row_idx, column=2, value=self._format_value(value)).border = border
            
    def _write_list_to_sheet(self, ws, data, header_font, header_fill, border):
        """Write list to Excel sheet"""
        if not data:
            return
            
        if isinstance(data[0], dict):
            # List of dicts
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                
            for row_idx, item in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=self._format_value(item.get(header)))
                    cell.border = border
        else:
            # Simple list
            for row_idx, item in enumerate(data, 1):
                ws.cell(row=row_idx, column=1, value=self._format_value(item)).border = border
                
    def _format_value(self, value):
        """Format value for Excel"""
        if pd.isna(value):
            return ''
        if isinstance(value, (datetime, pd.Timestamp)):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, (list, dict)):
            return json.dumps(value)
        return value
    
    def export_to_pdf(self, data: Dict[str, Any], filename: str = None,
                     title: str = "Financial Dashboard Report") -> Union[bytes, str]:
        """
        Export data to PDF format
        
        Args:
            data: Dictionary with sections and their data
            filename: If provided, save to file; otherwise return bytes
            title: Report title
            
        Returns:
            Bytes of PDF file or filepath
        """
        if not HAS_REPORTLAB:
            raise ImportError("reportlab required for PDF export. Install with: pip install reportlab")
            
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter,
                               rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=72)
        
        story = []
        
        # Title
        story.append(Paragraph(title, self.styles['CustomTitle']))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                              self.styles['CustomBody']))
        story.append(Spacer(1, 20))
        
        # Process each section
        for section_name, section_data in data.items():
            story.append(Paragraph(section_name, self.styles['SectionTitle']))
            
            if isinstance(section_data, pd.DataFrame):
                story.extend(self._df_to_pdf_table(section_data))
            elif isinstance(section_data, dict):
                story.extend(self._dict_to_pdf(section_data))
            elif isinstance(section_data, list):
                story.extend(self._list_to_pdf(section_data))
            elif isinstance(section_data, str):
                story.append(Paragraph(section_data, self.styles['CustomBody']))
                
            story.append(Spacer(1, 20))
        
        doc.build(story)
        output.seek(0)
        
        if filename:
            with open(filename, 'wb') as f:
                f.write(output.getvalue())
            return filename
        
        return output.getvalue()
    
    def _df_to_pdf_table(self, df: pd.DataFrame, max_rows: int = 50) -> List:
        """Convert DataFrame to PDF table"""
        elements = []
        
        # Limit rows
        if len(df) > max_rows:
            df = df.head(max_rows)
            elements.append(Paragraph(f"(Showing first {max_rows} rows)", self.styles['CustomBody']))
        
        # Prepare data
        table_data = [df.columns.tolist()]
        for _, row in df.iterrows():
            table_data.append([self._format_cell(v) for v in row.values])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')])
        ]))
        
        elements.append(table)
        return elements
    
    def _dict_to_pdf(self, data: dict) -> List:
        """Convert dictionary to PDF elements"""
        elements = []
        
        table_data = [['Key', 'Value']]
        for key, value in data.items():
            table_data.append([str(key), self._format_cell(value)])
        
        table = Table(table_data, colWidths=[2*inch, 4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        
        elements.append(table)
        return elements
    
    def _list_to_pdf(self, data: list) -> List:
        """Convert list to PDF elements"""
        elements = []
        
        if not data:
            return elements
            
        if isinstance(data[0], dict):
            df = pd.DataFrame(data)
            return self._df_to_pdf_table(df)
        else:
            for item in data:
                elements.append(Paragraph(f"• {self._format_cell(item)}", self.styles['CustomBody']))
                
        return elements
    
    def _format_cell(self, value) -> str:
        """Format cell value for PDF"""
        if pd.isna(value):
            return '-'
        if isinstance(value, float):
            if abs(value) < 0.01:
                return f'{value:.4f}'
            return f'{value:,.2f}'
        if isinstance(value, (datetime, pd.Timestamp)):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, (list, dict)):
            return json.dumps(value)[:50]
        return str(value)[:100]
    
    def export_portfolio_report(self, portfolio_data: Dict[str, Any],
                               positions: pd.DataFrame,
                               performance: pd.DataFrame,
                               format: str = 'excel') -> bytes:
        """Generate comprehensive portfolio report"""
        data = {
            'Summary': pd.DataFrame([portfolio_data]),
            'Positions': positions,
            'Performance': performance
        }
        
        if format == 'excel':
            return self.export_to_excel(data)
        else:
            return self.export_to_pdf(data, title="Portfolio Report")
    
    def export_options_analysis(self, ticker: str,
                               chain_data: pd.DataFrame,
                               greeks: Dict[str, float],
                               iv_data: Dict[str, Any],
                               format: str = 'excel') -> bytes:
        """Export options analysis report"""
        data = {
            'Options Chain': chain_data,
            'Greeks': greeks,
            'IV Analysis': iv_data
        }
        
        if format == 'excel':
            return self.export_to_excel(data)
        else:
            return self.export_to_pdf(data, title=f"Options Analysis - {ticker}")
    
    def export_risk_report(self, risk_data: Dict[str, Any],
                          format: str = 'pdf') -> bytes:
        """Export risk analysis report"""
        if format == 'excel':
            return self.export_to_excel(risk_data)
        else:
            return self.export_to_pdf(risk_data, title="Risk Analysis Report")


# Singleton instance
_export_service = None

def get_export_service() -> ExportService:
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service


def create_download_link(data: bytes, filename: str, format: str = 'xlsx') -> str:
    """Create download link for Dash/HTML"""
    b64 = base64.b64encode(data).decode()
    
    if format == 'xlsx':
        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif format == 'pdf':
        mime = 'application/pdf'
    else:
        mime = 'application/octet-stream'
    
    return f'data:{mime};base64,{b64}'
